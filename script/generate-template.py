#!/usr/bin/env python3
"""Regenerates client/public/showflow-template.xlsx (the annotated template the app serves).

Run from the repo root:  python3 script/generate-template.py
Requires: openpyxl
"""
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

OUT = "client/public/showflow-template.xlsx"

REQUIRED_FILL = PatternFill("solid", fgColor="1F2937")
SPECIAL_FILL = PatternFill("solid", fgColor="B45309")  # amber — special optional column
OPTIONAL_FILL = PatternFill("solid", fgColor="6B7280")
HEADER_FONT = Font(bold=True, color="FFFFFF")
COLOUR_FILLS = {"Yellow": "FFCC78", "Green": "77C785", "Purple": "A790F5", "Blue": "779BE7"}

# (header, width, fill, comment)
COLUMNS = [
    ("Cue #", 7, REQUIRED_FILL, "REQUIRED\nSequential integer cue number: 1, 2, 3… No letters or gaps."),
    ("Start Time", 11, REQUIRED_FILL,
     "REQUIRED\nClock time the item starts, e.g. 9:00 AM or 13:45.\n"
     "Only needs to be exact when Linkstart is FALSE (pinned rows); linked rows are recalculated from the previous item."),
    ("Duration", 10, REQUIRED_FILL, "REQUIRED\nHow long the item runs, e.g. 0:05:00 or 5:00 (h:mm:ss or m:ss)."),
    ("End Time", 11, REQUIRED_FILL,
     "REQUIRED\nClock time the item ends (Start Time + Duration).\nTip: use a formula so it always matches."),
    ("Linkstart", 10, REQUIRED_FILL,
     "REQUIRED\nTRUE  = this item starts the moment the previous item ends (start time follows automatically).\n"
     "FALSE = this item is pinned to the clock time in Start Time (use for doors, breaks, anything that must hit an exact time)."),
    ("Title", 28, REQUIRED_FILL,
     "REQUIRED\nWhat appears on the Ontime rundown, e.g. \"Walk-in\", \"CEO Welcome + PPT\".\n"
     "Avoid the word \"keynote\" — it will be replaced with \"PPT\"."),
    ("Timer Type", 11, REQUIRED_FILL,
     "REQUIRED (may be omitted entirely)\n"
     "Controls the item's primary timer in Ontime:\n"
     "none        = no countdown timer for this item\n"
     "count-down  = show a countdown timer (typical for speaker segments)\n"
     "If you delete this whole column, the timer type is inferred from Colour instead: Blue → count-down, everything else → none."),
    ("Colour", 9, REQUIRED_FILL,
     "REQUIRED\nOne of: Yellow, Green, Purple, Blue\n"
     "Yellow = videos · Green = breaks / walk-ins · Purple = intros / outros · Blue = speaking segments"),
    ("Aux Timer", 10, SPECIAL_FILL,
     "OPTIONAL — special column (NOT a custom field).\n"
     "Drives Ontime's Aux timer 1 with automations built at sync time.\n"
     "Put a duration (e.g. 1:00:00 or 0:40:00) on the row where the aux timer should reset and start counting down. "
     "It keeps running across items until another row resets it, and it stops when you stop the show.\n"
     "00:00:00 clears the timer: it stops, zeroes out, and displays show nothing until the next reset row.\n"
     "Leave cells blank or write \"none\" on rows that don't touch it — or delete the whole column if you don't use aux timers.\n"
     "Keep Aux timer 1 set to count-down direction in Ontime."),
    ("Screenstate", 12, OPTIONAL_FILL,
     "OPTIONAL — becomes a custom field in Ontime.\nWhich screen look is live, e.g. SS1 / SS2 / SS3."),
    ("Video", 18, OPTIONAL_FILL, "OPTIONAL — becomes a custom field in Ontime.\nVideo cue for this item, e.g. \"Roll opener video\"."),
    ("Lighting", 16, OPTIONAL_FILL, "OPTIONAL — becomes a custom field in Ontime.\nLighting cue, e.g. \"Stage wash\", \"House to half\"."),
    ("Audio", 17, OPTIONAL_FILL, "OPTIONAL — becomes a custom field in Ontime.\nAudio cue, e.g. \"Podium mic live\", \"Play walk-in music\"."),
    ("Speakers", 10, OPTIONAL_FILL, "OPTIONAL — becomes a custom field in Ontime.\nWho is on stage / at the podium."),
    ("Stage", 12, OPTIONAL_FILL, "OPTIONAL — becomes a custom field in Ontime.\nStage action, e.g. \"Lectern on\", \"Panel chairs set\"."),
    ("Notes", 30, OPTIONAL_FILL,
     "OPTIONAL\nFree-form production notes (goes to the Ontime note field, not a custom field).\n\n"
     "Add as many extra columns as you like after the required eight — every extra column is created automatically as an Ontime custom field."),
]

ROWS = [
    [1, "9:00 AM", "0:30:00", "9:30 AM", "FALSE", "Walk-in — doors open", "none", "Green", "", "SS1", "", "House look", "Walk-in music", "", "", ""],
    [2, "9:30 AM", "0:01:00", "9:31 AM", "TRUE", "Opener video", "none", "Yellow", "", "SS1", "Roll opener video", "Stage dark", "Video audio full", "", "", "Video from client"],
    [3, "9:31 AM", "0:02:00", "9:33 AM", "TRUE", "MC intro", "none", "Purple", "", "SS2", "", "Stage wash", "Podium mic live", "MC", "", "MC walks from SR"],
    [4, "9:33 AM", "0:25:00", "9:58 AM", "TRUE", "CEO Welcome + PPT", "count-down", "Blue", "1:00:00", "SS3", "Advance PPT", "Podium special", "Podium mic", "CEO", "Lectern on", "Aux resets to 1h here and counts down"],
    [5, "9:58 AM", "0:02:00", "10:00 AM", "TRUE", "MC outro to break", "none", "Purple", "none", "SS2", "", "Stage wash", "Podium mic live", "MC", "", ""],
    [6, "10:00 AM", "0:15:00", "10:15 AM", "FALSE", "Break", "none", "Green", "0:15:00", "SS1", "", "House look", "Break music", "", "", "Pinned — break always starts at 10:00 · aux resets to 15 min"],
]

HOW_IT_WORKS = [
    (1, "Mosc-tools — Ontime Show Flow Sync: sheet template", True, 12),
    (3, "1.  Keep the first eight columns exactly as named — they are required:", True, 11),
    (4, "     Cue #, Start Time, Duration, End Time, Linkstart, Title, Timer Type, Colour", False, 11),
    (6, "2.  Linkstart controls timing:", True, 11),
    (7, "     TRUE  → the item starts when the previous one ends (rolls with the show)", False, 11),
    (8, "     FALSE → the item is pinned to the clock time in Start Time (doors, breaks, hard outs)", False, 11),
    (10, "3.  Timer Type drives each item's primary timer: 'count-down' shows a countdown in Ontime, 'none' does not.", True, 11),
    (11, "     You may delete the whole column — the timer type is then inferred from Colour (Blue → count-down, others → none).", False, 11),
    (13, "4.  Colour must be Yellow, Green, Purple or Blue:", True, 11),
    (14, "     Yellow = videos · Green = breaks / walk-ins · Purple = intros / outros · Blue = speaking segments", False, 11),
    (16, "5.  Aux Timer (optional, special): drives Ontime's Aux timer 1 via automations created at sync time.", True, 11),
    (17, "     Put a duration (e.g. 1:00:00) on the row where the aux timer should reset and start counting down.", False, 11),
    (18, "     It keeps running across items until another row resets it, and it stops when you stop the show.", False, 11),
    (19, "     00:00:00 clears the timer — it stops, zeroes out, and displays show nothing until the next reset row.", False, 11),
    (20, "     Blank or 'none' on rows that don't touch it. Delete the column if you don't use aux timers.", False, 11),
    (21, "6.  Every other column you add after the required eight becomes an Ontime custom field automatically", True, 11),
    (22, "     (Screenstate, Video, Lighting, Audio, Speakers, Stage are just examples — rename or add your own).", False, 11),
    (23, "     'Notes' is special: it fills the Ontime note field instead.", False, 11),
    (25, "7.  Upload this file to Google Sheets (File → Import), then paste the sheet ID and tab name", True, 11),
    (26, "     into the tool's Settings (gear icon).", False, 11),
    (28, "Hover over any header cell on the 'Show Flow' tab for details on that column.", False, 11),
    (29, "Technical support: mosc-tools@moscone.ca", False, 11),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Show Flow"
    ws.freeze_panes = "A2"

    for i, (header, width, fill, note) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.fill = fill
        cell.font = HEADER_FONT
        ws.column_dimensions[cell.column_letter].width = width
        comment = Comment(note, "Mosc-tools", height=180, width=340)
        cell.comment = comment

    for r, row in enumerate(ROWS, start=2):
        for c, value in enumerate(row, start=1):
            if value == "":
                continue
            cell = ws.cell(row=r, column=c, value=value)
            if COLUMNS[c - 1][0] == "Colour" and value in COLOUR_FILLS:
                cell.fill = PatternFill("solid", fgColor=COLOUR_FILLS[value])

    ws2 = wb.create_sheet("How it works")
    ws2.column_dimensions["A"].width = 110
    for row, text, bold, size in HOW_IT_WORKS:
        cell = ws2.cell(row=row, column=1, value=text)
        cell.font = Font(bold=bold, size=size)
        cell.alignment = Alignment(vertical="top")

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
